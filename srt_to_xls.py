#!/usr/bin/env python3

import pandas as pd
import re
import argparse
from datetime import timedelta
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import copy  # For handling style copying

def srt_to_smpte(timecode, frame_rate):
    """Convert SRT timecode to SMPTE timecode."""
    hours, minutes, seconds, milliseconds = map(int, re.split('[:,]', timecode))
    total_seconds = timedelta(hours=hours, minutes=minutes, seconds=seconds, milliseconds=milliseconds).total_seconds()
    frames = int((total_seconds * frame_rate) % frame_rate)
    return f"{hours:02}:{minutes:02}:{seconds:02}:{frames:02}"

def parse_srt(srt_file, frame_rate):
    """Parse the SRT file and extract line number, timecodes, script text, and notes."""
    with open(srt_file, 'r') as file:
        content = file.read()
    
    pattern = re.compile(r'(\d+)\n(\d{2}:\d{2}:\d{2},\d{3}) --> (\d{2}:\d{2}:\d{2},\d{3})\n(.*?)\n\n', re.DOTALL)
    matches = pattern.findall(content)

    data = []
    for match in matches:
        line_number = int(match[0])
        timecode_in = srt_to_smpte(match[1], frame_rate)
        timecode_out = srt_to_smpte(match[2], frame_rate)
        script_text = match[3].replace('\n', ' ')
        
        # Extract content in square brackets and modify the line
        bracket_content = re.search(r'\[(.*?)\]', script_text)
        if bracket_content:
            # Ensure no extra spaces appear around the extracted note
            note = f"[{bracket_content.group(1).strip()}]"
            # Remove the brackets and their content from the main text
            script_text = re.sub(r'\[.*?\]', '', script_text).strip()
        else:
            note = None

        data.append([line_number, timecode_in, timecode_out, script_text, note])

    return data

def copy_style(source_cell, target_cell):
    """Copy style from one cell to another using updated methods."""
    if source_cell.font:
        target_cell.font = copy.copy(source_cell.font)
    if source_cell.border:
        target_cell.border = copy.copy(source_cell.border)
    if source_cell.fill:
        target_cell.fill = copy.copy(source_cell.fill)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.protection:
        target_cell.protection = copy.copy(source_cell.protection)
    if source_cell.alignment:
        target_cell.alignment = copy.copy(source_cell.alignment)

def srt_to_excel(srt_file, excel_file, frame_rate, template_file=None):
    """Convert the parsed SRT data into an Excel file, optionally using a template."""
    data = parse_srt(srt_file, frame_rate)
    df = pd.DataFrame(data, columns=['Line Number', 'Timecode In', 'Timecode Out', 'Script (en)', 'On Screen Note (en)'])
    
    # Ensure the output file has the .xlsx extension
    if not excel_file.lower().endswith('.xlsx'):
        excel_file += '.xlsx'

    if template_file:
        wb = openpyxl.load_workbook(template_file)
        ws = wb.active

        # Find the last row with formatting
        last_row_with_format = ws.max_row

        # Start writing data from the second row
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx <= last_row_with_format:
                    # Copy formatting from the last formatted row
                    template_cell = ws.cell(row=last_row_with_format, column=c_idx)
                    copy_style(template_cell, cell)

        wb.save(excel_file)
    else:
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=1)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Convert SRT file to Excel AD script.')
    parser.add_argument('srt_file', help='Path to the SRT file')
    parser.add_argument('excel_file', nargs='?', help='Path to the output Excel file (optional, defaults to the SRT file name ending with _studioscript.xlsx)')
    parser.add_argument('frame_rate', type=float, help='Frame rate of the video (e.g., 23.976, 24, 25, 30)')
    parser.add_argument('--template', help='Path to the Excel template file', default=None)
    args = parser.parse_args()

    # Default template file path
    DEFAULT_TEMPLATE_PATH = "~/Documents/studioscript_template.xlsx"
    
    # Use the default template if none is provided
    if not args.template:
        args.template = os.path.expanduser(DEFAULT_TEMPLATE_PATH)

    # Default Excel file name if not provided
    if not args.excel_file:
        base_name = os.path.splitext(os.path.basename(args.srt_file))[0]
        args.excel_file = base_name + "_studioscript.xlsx"

    srt_to_excel(args.srt_file, args.excel_file, args.frame_rate, args.template)
