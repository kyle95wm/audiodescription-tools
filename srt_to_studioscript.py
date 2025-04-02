#!/usr/bin/env python3

import pandas as pd
import re
import argparse
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def normalize_frame_rate(fps):
    """Convert common shorthand frame rates to exact float equivalents."""
    known_rates = {
        23.976: 24000 / 1001,
        29.97: 30000 / 1001,
        59.94: 60000 / 1001,
    }
    return known_rates.get(round(fps, 3), fps)

def srt_to_smpte(timecode, frame_rate):
    """Convert SRT timecode to SMPTE timecode.

    Frame count is based on real frame rate (e.g., 23.976),
    but SMPTE display is aligned to runtime (e.g., 24 fps).
    """
    hours, minutes, seconds, milliseconds = map(int, re.split('[:,]', timecode))
    total_seconds = hours * 3600 + minutes * 60 + seconds + milliseconds / 1000

    # Real total frames based on actual frame rate
    total_frames = total_seconds * frame_rate

    # Convert to SMPTE as if 24 fps
    smpte_hours = int(total_frames // (3600 * 24))
    total_frames %= (3600 * 24)

    smpte_minutes = int(total_frames // (60 * 24))
    total_frames %= (60 * 24)

    smpte_seconds = int(total_frames // 24)
    smpte_frames = int(total_frames % 24)

    return f"{smpte_hours:02}:{smpte_minutes:02}:{smpte_seconds:02}:{smpte_frames:02}"

def parse_srt(srt_file, frame_rate):
    """Parse the SRT file and extract line number, timecodes, script text, and notes."""
    with open(srt_file, 'r') as file:
        content = file.read()

    pattern = re.compile(r'(\d+)\n(\d{2}:\d{2}:\d{2},\d{3}) --> (\d{2}:\d{2}:\d{2},\d{3})\n(.*?)\n\n', re.DOTALL)
    matches = pattern.findall(content)

    data = []
    for match in matches:
        line_number = int(match[0])
        timecode_in = srt_to_smpte(match[1], frame_rate)
        timecode_out = srt_to_smpte(match[2], frame_rate)
        script_text = match[3].replace('\n', ' ')

        bracket_content = re.search(r'\[(.*?)\]', script_text)
        if bracket_content:
            note = f"[{bracket_content.group(1).strip()}]"
            script_text = re.sub(r'\[.*?\]', '', script_text).strip()
        else:
            note = None

        data.append([line_number, timecode_in, timecode_out, script_text, note])

    return data

def create_default_template():
    """Create a default Excel workbook that matches the provided template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Studio Script"

    headers = ["Line Number", "Timecode In", "Timecode Out", "Script (en)", "On Screen Note (en)"]
    column_widths = [12, 15, 15, 50, 30]

    for col_num, (header, width) in enumerate(zip(headers, column_widths), start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    return wb

def srt_to_excel(srt_file, excel_file, frame_rate, template_file=None):
    """Convert the parsed SRT data into an Excel file, optionally using a template."""
    data = parse_srt(srt_file, frame_rate)
    df = pd.DataFrame(data, columns=['Line Number', 'Timecode In', 'Timecode Out', 'Script (en)', 'On Screen Note (en)'])

    if not excel_file.lower().endswith('.xlsx'):
        excel_file += '.xlsx'

    if template_file and os.path.exists(template_file):
        wb = openpyxl.load_workbook(template_file)
        ws = wb.active
    else:
        print(f"Warning: Template file '{template_file}' not found. Using default template.")
        wb = create_default_template()
        ws = wb.active

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(size=16)
            if isinstance(value, str):
                cell.alignment = Alignment(wrap_text=True)

    for row_dim in ws.row_dimensions.values():
        row_dim.height = None

    wb.save(excel_file)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Convert SRT file to Excel AD script.')
    parser.add_argument('srt_file', help='Path to the SRT file')
    parser.add_argument('excel_file', nargs='?', help='Path to the output Excel file (optional)')
    parser.add_argument('frame_rate', type=float, help='Frame rate of the video (e.g., 23.976, 24, 25, 30)')
    parser.add_argument('--template', help='Path to the Excel template file', default=None)
    args = parser.parse_args()

    args.frame_rate = normalize_frame_rate(args.frame_rate)

    if not args.template:
        args.template = os.path.expanduser("~/Documents/studioscript_template.xlsx")

    if not args.excel_file:
        base_name = os.path.splitext(os.path.basename(args.srt_file))[0]
        args.excel_file = base_name + "_studioscript.xlsx"

    srt_to_excel(args.srt_file, args.excel_file, args.frame_rate, args.template)
