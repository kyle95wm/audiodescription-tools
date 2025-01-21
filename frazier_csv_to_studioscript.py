#!/usr/bin/env python3

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse
import os
import re

def extract_dialogue_notes(dialogue):
    """Extract content from dialogue, retaining square brackets."""
    if pd.isna(dialogue) or dialogue == "0":
        return None  # Skip empty or placeholder fields
    match = re.search(r'\[(.*?)\]', dialogue)  # Match content in square brackets
    return f"[{match.group(1)}]" if match else dialogue  # Retain brackets or return full dialogue

def create_default_template():
    """Create a default Excel workbook with headers and formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Studio Script"

    # Define headers and column widths
    headers = ["Line Number", "Timecode In", "Timecode Out", "Script/Text", "Dialogue (Notes)"]
    column_widths = [12, 15, 15, 50, 30]

    for col_num, (header, width) in enumerate(zip(headers, column_widths), start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    return wb

def csv_to_excel(input_csv, output_excel, template_path=None):
    """Convert CSV to Excel studio script format using a template or default layout."""
    # Read the CSV file
    csv_data = pd.read_csv(input_csv, sep=";")

    # Filter and rename columns
    csv_data["Dialogue"] = csv_data["Dialogue"].apply(extract_dialogue_notes)  # Process "Dialogue" field
    filtered_data = csv_data[["Position", "Start", "End", "Text", "Dialogue"]]
    filtered_data.columns = ["Line Number", "Timecode In", "Timecode Out", "Script/Text", "Dialogue (Notes)"]

    # Load or create a workbook
    if template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        print(f"Using template file: {template_path}")
    else:
        print(f"Template not found. Using default layout.")
        wb = create_default_template()
        ws = wb.active

    # Write data into the worksheet, starting after the header row
    for row_idx, row in enumerate(dataframe_to_rows(filtered_data, index=False, header=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, str):  # Enable text wrapping for strings
                cell.alignment = Alignment(wrap_text=True)

    # Save the workbook
    wb.save(output_excel)
    print(f"Excel file saved to: {output_excel}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert CSV to Studio Script Excel format")
    parser.add_argument("input_csv", help="Path to the input CSV file")
    parser.add_argument(
        "output_excel", nargs="?", help="Path to the output Excel file (optional, defaults to appending '_studioscript.xlsx')"
    )
    parser.add_argument("--template", help="Path to the Excel template file", default=None)
    args = parser.parse_args()

    # Determine the default output file name if not provided
    if not args.output_excel:
        base_name = os.path.splitext(os.path.basename(args.input_csv))[0]
        args.output_excel = f"{base_name}_studioscript.xlsx"

    # Default template path (if none provided)
    DEFAULT_TEMPLATE_PATH = "~/Documents/studioscript_template.xlsx"
    if not args.template:
        args.template = os.path.expanduser(DEFAULT_TEMPLATE_PATH)

    # Run the conversion
    csv_to_excel(args.input_csv, args.output_excel, args.template)
